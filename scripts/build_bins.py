#!/usr/bin/env python3
"""서울시 가로쓰레기통 설치정보 → 좌표 부여(지오코딩) → bins.json.

쓰레기통 자료에는 좌표 없이 주소("사직로 125")만 있다. 주소→좌표는 저장소에
보존된 사전으로 해결한다. 외부 지오코딩 API 불필요.

매칭 규칙 (순서대로 시도):
  1. 확정 좌표 사전(data/cache/addr_coords.json — 이전 빌드에서 확정한 주소↔좌표)
  2. Nominatim 캐시(data/cache/nominatim_bins.json — geocode_bins.py 가 생성)
  3. 실패 → unmatched (bins.json에 미포함, 개수만 기록)

입력: data/raw/bins_202511.xlsx
출력: docs/data/bins.json
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data/raw"
OUT = ROOT / "docs/data/bins.json"


def load_caches():
    """(확정 좌표 사전, Nominatim 캐시) — 각각 "구 주소" -> [lon, lat]. 없으면 빈 dict."""
    addr_path = ROOT / "data/cache/addr_coords.json"
    addrcache = json.loads(addr_path.read_text(encoding="utf-8")) if addr_path.exists() else {}
    cache_path = ROOT / "data/cache/nominatim_bins.json"
    geocache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else {}
    return addrcache, geocache


def lookup(addrcache, geocache, key):
    """확정 사전 → Nominatim 캐시 순으로 좌표를 찾는다. 없거나 null 이면 None."""
    g = addrcache.get(key) or geocache.get(key)
    return (g[0], g[1]) if g else None


def main() -> None:
    addrcache, geocache = load_caches()
    wb = openpyxl.load_workbook(RAW / "bins_202511.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    bins, unmatched = [], []
    methods = defaultdict(int)
    for row in ws.iter_rows(min_row=6, values_only=True):
        _no, gu, addr, detail, place, kind = (list(row) + [None] * 6)[:6]
        if not gu or not addr:
            continue
        gu, addr = str(gu).strip(), str(addr).strip()
        g = lookup(addrcache, geocache, f"{gu} {addr}")
        if g is None:
            unmatched.append(f"{gu} {addr}")
            continue
        lon, lat = g
        methods["cache"] += 1
        bins.append({
            "lon": round(lon, 6), "lat": round(lat, 6),
            "gu": gu, "addr": addr,
            "d": str(detail or "").strip(), "k": str(kind or "").strip(),
        })
    OUT.write_text(json.dumps({
        "asof": "2025-11", "total": len(bins) + len(unmatched),
        "matched": len(bins), "sensors": None, "bins": bins,
    }, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"total={len(bins)+len(unmatched)} matched={len(bins)} "
          f"({dict(methods)}) unmatched={len(unmatched)}")
    for u in unmatched[:10]:
        print("  miss:", u)


if __name__ == "__main__":
    sys.exit(main())
