#!/usr/bin/env python3
"""확정 좌표 사전·Nominatim 캐시로도 좌표를 못 찾은 쓰레기통 주소를 Nominatim(OSM)으로 지오코딩.

- 결과는 data/cache/nominatim_bins.json 에 누적 (주소 -> [lon,lat] 또는 null)
- 초당 1건 제한(Nominatim 이용정책) 준수. 중단 후 재실행하면 이어서 진행
- 뭉툭한 결과(동·구 중심점 등)는 실패로 취급: place/quarter/suburb/... 제외
- 완료 후 build_bins.py 를 다시 실행하면 캐시가 반영된 bins.json 이 나온다
"""
import json
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from build_bins import RAW, load_caches, lookup  # noqa: E402

import openpyxl

CACHE = Path(__file__).resolve().parent.parent / "data/cache/nominatim_bins.json"
UA = {"User-Agent": "golmok-daejang geocoder (github.com/grboy6770/golmok-daejang)"}
REJECT_TYPES = {"quarter", "neighbourhood", "suburb", "borough", "city", "county", "state", "postcode"}


def query(addr: str):
    q = urllib.parse.quote("서울특별시 " + addr)
    url = f"https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=1&countrycodes=kr"
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=20) as r:
        d = json.load(r)
    if not d:
        return None
    top = d[0]
    if top.get("type") in REJECT_TYPES or top.get("class") == "boundary":
        return None
    lon, lat = float(top["lon"]), float(top["lat"])
    if not (126.7 <= lon <= 127.3 and 37.4 <= lat <= 37.8):
        return None
    return [round(lon, 6), round(lat, 6)]


def main() -> None:
    addrcache, geocache = load_caches()
    wb = openpyxl.load_workbook(RAW / "bins_202511.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    misses = []
    seen = set()
    for row in ws.iter_rows(min_row=6, values_only=True):
        _no, gu, addr = (list(row) + [None] * 3)[:3]
        if not gu or not addr:
            continue
        gu, addr = str(gu).strip(), str(addr).strip()
        key = f"{gu} {addr}"
        if key in seen:
            continue
        seen.add(key)
        if lookup(addrcache, geocache, key) is None:
            misses.append(key)

    CACHE.parent.mkdir(parents=True, exist_ok=True)
    cache = json.loads(CACHE.read_text(encoding="utf-8")) if CACHE.exists() else {}
    todo = [m for m in misses if m not in cache]
    print(f"고유 미매칭 {len(misses)}건, 캐시 제외 조회 대상 {len(todo)}건")
    for i, key in enumerate(todo, 1):
        try:
            cache[key] = query(key)
        except Exception as e:
            print(f"  중단({e}) — 진행분은 저장됨", file=sys.stderr)
            break
        if i % 50 == 0 or i == len(todo):
            CACHE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
            hit = sum(1 for k in todo[:i] if cache.get(k))
            print(f"  {i}/{len(todo)} (적중 {hit})")
        time.sleep(1.05)
    CACHE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    hits = sum(1 for m in misses if cache.get(m))
    print(f"완료: 미매칭 {len(misses)}건 중 지오코딩 성공 {hits}건")


if __name__ == "__main__":
    sys.exit(main())
