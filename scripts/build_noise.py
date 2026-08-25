#!/usr/bin/env python3
"""S-DoT 주간 환경정보 CSV → 센서별 시간대 평균 소음 JSON.

입력 (data/raw/):
  - sdot_loc.xlsx : 센서 설치 위치 (시리얼, 주소, 위도, 경도, 변경 전 시리얼)
  - sdot_*.csv    : 주간 환경정보 (cp949). "소음 평균(dB)" 컬럼 사용
출력:
  - docs/data/noise.json : {"weeks": [...], "sensors": [{"s","lon","lat","addr","h":[24개 시간대 평균 dB]}]}
    h[i]는 i시(현지) 평균 소음. 해당 시간 측정이 없으면 -1.
    (null을 쓰면 MapLibre 색 계산식이 평가 오류로 기본색을 그려서 -1 표기로 통일)
"""
import csv
import glob
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data/raw"
OUT = ROOT / "docs/data/noise.json"


def load_locations() -> dict:
    wb = openpyxl.load_workbook(RAW / "sdot_loc.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    loc = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        no, serial, addr, _code, lat, lon, old1, old2 = (list(row) + [None] * 8)[:8]
        if not serial or lat is None or lon is None:
            continue
        entry = {"addr": str(addr or "").replace("서울특별시 ", ""), "lat": float(lat), "lon": float(lon)}
        loc[str(serial).strip()] = entry
        for old in (old1, old2):
            if old and str(old).strip():
                loc.setdefault(str(old).strip(), entry)
    return loc


def main() -> None:
    loc = load_locations()
    acc: dict[str, list[list[float]]] = {}  # serial -> [[sum,count] x 24]
    files = sorted(glob.glob(str(RAW / "sdot_*.csv")))
    unmatched = set()

    for path in files:
        with open(path, encoding="cp949", newline="") as f:
            reader = csv.reader(f)
            header = next(reader)
            i_serial = header.index("시리얼")
            i_time = header.index("측정시간")
            i_noise = header.index("소음 평균(dB)")
            for row in reader:
                if len(row) <= i_noise:
                    continue
                val = row[i_noise].strip()
                if not val:
                    continue
                try:
                    db = float(val)
                except ValueError:
                    continue
                if not (20 <= db <= 120):  # 센서 오류값 제외
                    continue
                t = row[i_time]  # 2026-08-10_00:07:00
                try:
                    hour = int(t.split("_")[1][:2])
                except (IndexError, ValueError):
                    continue
                serial = row[i_serial].strip()
                if serial not in loc:
                    unmatched.add(serial)
                    continue
                slot = acc.setdefault(serial, [[0.0, 0] for _ in range(24)])
                slot[hour][0] += db
                slot[hour][1] += 1

    sensors = []
    for serial, slots in acc.items():
        h = [round(s / c, 1) if c else -1 for s, c in slots]
        if all(v < 0 for v in h):
            continue
        e = loc[serial]
        sensors.append({"s": serial, "lon": round(e["lon"], 6), "lat": round(e["lat"], 6),
                        "addr": e["addr"], "h": h})

    OUT.write_text(json.dumps({
        "weeks": [Path(p).stem for p in files],
        "sensors": sensors,
    }, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"files={len(files)} sensors_with_noise={len(sensors)} unmatched_serials={len(unmatched)}")


if __name__ == "__main__":
    sys.exit(main())
